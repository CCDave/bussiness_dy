#!/usr/bin/env python
# coding: utf-8
from pickle import TRUE
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import numpy as np
import warnings
warnings.filterwarnings("ignore")

config = {'订单表名': '订单表.csv',
          '售后表名': '售后表.xlsx',
          '单品退货率': {
              '商品ID': [3568637652754402212],
              '订单日期': ['2022-09-13']
          },
          '单品发货率': {
              '商品ID': [3568637652754402212],
              '订单日期': ['2022-09-13']
          },
          '所有订单数据': True
          }

orders_from_file = {}
orders_return_from_file = {}


# ## 通过groupby 获取数组 返回数组 或 0


def safe_get_group(g, key):
    if key in g.groups:
        return g.get_group(key)
    return []


def safe_get_group_len(groups, key):
    return len(safe_get_group(groups, key))


def write_excel_center(file_path, rang='A1:AE15'):
    wb = load_workbook(file_path)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        alignment_center = Alignment(horizontal='center', vertical='center')
        ws_area = ws[rang]
        for i in ws_area:
            for j in i:
                j.alignment = alignment_center
    wb.save(file_path)


def write_excel_file(data, file_path, sheet_name):
    data.to_excel(file_path, sheet_name=sheet_name)
    write_excel_center(file_path)

# ## 计算订单的整体退货率


def count_product_return_data(return_orders,  orders):

    orders_groups = orders.groupby('订单提交日期')
    return_orders_groups = return_orders.groupby('订单提交日期')
    columns = []
    for name, group in orders_groups:
        columns.append(name.strftime('%m-%d'))
    # 创建空表
    ret = pd.DataFrame(data=None, columns=columns,
                       index=['当天订单量',
                              '当天未发货退货数', '当天未发货退货率',
                              '当天发货退货数', '当天发货退货率',
                              '当天整体退货数', '当天整体退货率',
                              #'当天订单发货量', '当天订单发货率',
                              '累计订单量',
                              '累计未发货退货数', '累计未发货退货率',
                              '累计发货退货数', '累计发货退货率',
                              '累计整体退货数', '累计整体退货率'
                              #'累计发货量', '累计发货率'
                              ])
    total_orders_count = 0
    total_no_send_returns = 0
    total_return_after_send = 0
    total_all_return = 0
    total_send_count = 0
    index = 0
    for name, group in orders_groups:

        current_orders_count = len(group)

        return_orders = safe_get_group(return_orders_groups, name)
        send_return = return_return = no_send_return = all_return = return_after_send = 0
        if len(return_orders) != 0 and current_orders_count != 0:
            send_return, return_return, no_send_return, all_return, return_after_send = count_orders_return_data(
                return_orders)
        current_no_send_returns = no_send_return
        current_return_after_send = return_after_send
        current_all_return = all_return

        total_orders_count = total_orders_count + current_orders_count
        total_no_send_returns = total_no_send_returns + current_no_send_returns
        total_return_after_send = total_return_after_send + current_return_after_send
        total_all_return = total_all_return + current_all_return

        ret.iloc[0][index] = current_orders_count
        ret.iloc[1][index] = current_no_send_returns
        ret.iloc[2][index] = "%.2f%%" % (
            current_no_send_returns/current_orders_count*100)

        ret.iloc[3][index] = current_return_after_send
        ret.iloc[4][index] = "%.2f%%" % (
            current_return_after_send/current_orders_count*100)

        ret.iloc[5][index] = current_all_return
        ret.iloc[6][index] = "%.2f%%" % (
            current_all_return/current_orders_count*100)

        ret.iloc[7][index] = total_orders_count

        ret.iloc[8][index] = total_no_send_returns
        ret.iloc[9][index] = "%.2f%%" % (
            total_no_send_returns/total_orders_count*100)

        ret.iloc[10][index] = total_return_after_send
        ret.iloc[11][index] = "%.2f%%" % (
            total_return_after_send/total_orders_count*100)

        ret.iloc[12][index] = total_all_return
        ret.iloc[13][index] = "%.2f%%" % (
            total_all_return/total_orders_count*100)
        index = index+1

    return ret


# ## 计算订单的退货数据细节


def count_orders_return_data(orders):
    return_type_group_type = orders.groupby('售后类型')
    send_return = safe_get_group_len(return_type_group_type, '已发货仅退款')
    return_return = safe_get_group_len(return_type_group_type, '退货退款')
    no_send_return = safe_get_group_len(return_type_group_type, '未发货仅退款')

    all_return = send_return + return_return + no_send_return
    return_after_send = send_return + return_return
    return send_return, return_return, no_send_return, all_return, return_after_send

# ## 计算订单的30天退货数据


def count_30days_return_data(orders, total_count):
    # 创建空表
    ret = pd.DataFrame(data=None, columns=['第1天', '第2天', '第3天',
                                           '第4天', '第5天', '第6天',
                                           '第7天', '第8天', '第9天',
                                           '第10天', '第11天', '第12天',
                                           '第13天', '第14天', '第15天',
                                           '第16天', '第17天', '第18天',
                                           '第19天', '第20天', '第21天',
                                           '第22天', '第23天', '第24天',
                                           '第25天', '第26天', '第27天',
                                           '第28天', '第29天', '第30天'],
                       index=['整体退货量', '整体退货率',
                              '发货前退货量', '发货前退货率',
                              '发货后退货量', '发货后退货率',
                              '总退货数累计', '总退货率累计'])

    orders_return_days_groups = orders.groupby('几天退款')
    days = 0
    cur_days_return_count = 0
    for name, group in orders_return_days_groups:
        days = name
        # print('第:%d天 退款单数:%d' % (name, len(group)))

        send_return, return_return, no_send_return, all_return, return_after_send = count_orders_return_data(
            group)

        # print('第:%d天, 退款单数:%d, 已发货仅退款：%d, 退货退款:%d,未发货仅退款:%d' %
        #      (name, len(group), send_return, return_return,  no_send_return))
        cur_days_return_count = cur_days_return_count + all_return

        ret.iloc[0][days] = all_return
        ret.iloc[1][days] = "%.2f%%" % (all_return/total_count*100)
        ret.iloc[2][days] = no_send_return
        ret.iloc[3][days] = "%.2f%%" % (no_send_return/total_count*100)
        ret.iloc[4][days] = return_after_send
        ret.iloc[5][days] = "%.2f%%" % (return_after_send/total_count*100)
        ret.iloc[6][days] = cur_days_return_count
        ret.iloc[7][days] = "%.2f%%" % (cur_days_return_count/total_count*100)

        if (days == 29):
            break
    return ret


def string_to_date(str):
    return datetime.date(*map(int, str.split('-')))


def get_product_name_from_table(table):
    return table.iloc[0]['选购商品'].strip()


if '订单表名' in config:
    orders_from_file = pd.read_csv(config['订单表名'])
else:
    print('！！！！！！！！订单表不存在,请检查文件 ！！！！！！！！')
    exit()

if '售后表名' in config:
    orders_return_from_file = pd.read_excel(config['售后表名'])
else:
    print('！！！！！！！！售后表不存在,请检查文件 ！！！！！！！！')
    exit()

if len(orders_from_file) == 0 or len(orders_return_from_file) == 0:
    print('订单表，或手表表数据存在问题。')

############################################# 初始化数据 ##################################################
# 提取订单提交日期：yyyy-mm-dd
orders_from_file['订单提交日期'] = pd.to_datetime(orders_from_file['订单提交时间']).dt.date
# 在退货表中新增两列‘主订单编号’&‘子订单编号’
orders_return_from_file['主订单编号'] = orders_return_from_file['订单号']
orders_return_from_file['子订单编号'] = orders_return_from_file['订单号']
# 提取售后申请日期：yyyy-mm-dd
orders_return_from_file['售后申请日期'] = pd.to_datetime(
    orders_return_from_file['售后申请时间']).dt.date
result = []
# 所有退款数据 算出第几天退款
# df_all是订单表和售后表和合集，主订单编号合并主键
df_all_return_orders = pd.merge(
    orders_return_from_file, orders_from_file, on='子订单编号', how='inner')
# print('合并订单....')
df_all_return_orders["几天退款"] = df_all_return_orders["售后申请日期"] - \
    df_all_return_orders["订单提交日期"]
# print('计算天数....')
# 转换成天数
df_all_return_orders["几天退款"] = df_all_return_orders["几天退款"].map(
    lambda x: x.days)
# df_all_return_orders.index = list(range(len(df_all_return_orders)))

# 从表按列的值筛选数据
# orders = orders_from_file[orders_from_file['支付完成时间'].notna()]
# 从表按列的值筛选数据
#orders = orders_from_file[orders_from_file['商品ID']==3568637652754402212]
#orders = orders_from_file[orders_from_file['支付完成时间']==pd.to_datetime('')]
#orders = orders_from_file[orders_from_file['主订单编号']==4977742859849892870]
# # 直接筛选方法
# orders = orders_from_file[(orders_from_file['支付完成时间'].notna()) &
#                        (orders_from_file['主订单编号'] == 4977752356016577699)]

# 计数
print(orders_from_file['主订单编号'].count())
print(orders_from_file['主订单编号'].nunique())

orders_from_file = pd.merge(
    orders_from_file, orders_return_from_file, on='子订单编号', how='inner')
orders_from_file.columns

print(orders_from_file.groupby(['商品ID_x', '订单提交日期','售后类型'])['主订单编号_x'].agg(
    [np.size]))
# print(orders)
