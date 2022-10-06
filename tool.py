#!/usr/bin/env python
# coding: utf-8
from pickle import TRUE
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import numpy as np
import warnings


def safe_get_group(g, key):
    if key in g.groups:
        return g.get_group(key)
    return []


def safe_get_group_len(groups, key):
    return len(safe_get_group(groups, key))


def write_excel_center(file_path, rang='A1:AE15'):
    wb = load_workbook(file_path)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        alignment_center = Alignment(horizontal='center', vertical='center')
        ws_area = ws[rang]
        for i in ws_area:
            for j in i:
                j.alignment = alignment_center
    wb.save(file_path)


def write_excel_file(data, file_path, sheet_name):
    data.to_excel(file_path, sheet_name=sheet_name)
    write_excel_center(file_path)
