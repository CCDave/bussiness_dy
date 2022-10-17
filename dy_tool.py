#!/usr/bin/env python
# coding: utf-8
from pickle import TRUE
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import numpy as np
import warnings


def safe_get_group(g, key):
    if key in g.groups:
        return g.get_group(key)
    return []


def safe_get_group_len(groups, key):
    return len(safe_get_group(groups, key))


def write_excel_center(file_path, rang='A1:AE15'):
    wb = load_workbook(file_path)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        alignment_center = Alignment(horizontal='center', vertical='center')
        ws_area = ws[rang]
        for i in ws_area:
            for j in i:
                j.alignment = alignment_center
    wb.save(file_path)


def write_excel_file(data, file_path, sheet_name):
    data.to_excel(file_path, sheet_name=sheet_name)
    write_excel_center(file_path)


def string_to_date(str):
    return datetime.date(*map(int, str.split('-')))


def get_product_name_from_table(table):
    return table.iloc[0]['选购商品'].strip()


def get_last_months(count):
    ret = []
    now_date_time = datetime.datetime.now()
    for i in range(0, count):
        ret.append(now_date_time.date)
        now_date_time = (now_date_time - datetime.timedelta(days=30))
        i = i + 1
    return ret


def currency_to_float(e):
    if isinstance(e, str):
        return e.replace(',', '')
    return e


def get_sum_size_from_status(orders, order_status, after_status):
    status = orders[(orders['订单状态'] == order_status) & (
        orders['售后状态'] == after_status)]
    if len(status):
        return status.iloc[0]['sum'], status.iloc[0]['size']
    else:
        return 0, 0
